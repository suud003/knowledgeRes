"""MCP Server - 为 Claude Code 提供个人助手工具.

支持功能:
- 数据同步 (sync_data)
- 内容收集 (collect_content) - 支持 Obsidian 格式和 __collected 目录
- 笔记添加 (add_note) - 支持 Obsidian 格式和 __created 目录
- 主题管理 (manage_topic) - 动态主题创建和管理
- Ship-Learn-Next 学习计划
"""

from __future__ import annotations

import asyncio
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from mcp.server.fastmcp import FastMCP

from pa.config import Config, load_config
from pa.config.settings import TopicConfig
from pa.formatters.obsidian import ObsidianFormatter
from pa.router.engine import RouterEngine
from pa.topics.manager import TopicManager

# 添加 ship-learn-next 到路径
SKILL_DIR = Path(__file__).parent.parent.parent / "skills" / "ship-learn-next" / "scripts"
if str(SKILL_DIR) not in sys.path:
    sys.path.insert(0, str(SKILL_DIR))

# 创建 MCP Server
mcp = FastMCP("personal-assistant")

# 全局配置缓存
_config: Config | None = None
_topic_manager: TopicManager | None = None


def get_config() -> Config:
    """获取配置（带缓存）."""
    global _config
    if _config is None:
        _config = load_config()
    return _config


def get_topic_manager() -> TopicManager:
    """获取主题管理器（带缓存）."""
    global _topic_manager
    if _topic_manager is None:
        config = get_config()
        # 将相对路径转换为绝对路径
        raw_dir = Path(config.sync.raw_dir).resolve()
        _topic_manager = TopicManager(config, raw_dir)
    return _topic_manager


@mcp.tool()
async def sync_data() -> str:
    """触发数据同步和路由，将所有数据源的数据同步到 Context 文件夹.

    Returns:
        同步结果摘要，包含处理的数据条数和写入的文件列表
    """
    try:
        config = get_config()
        from pa.collectors.feishu import FeishuCollector

        raw_dir = Path(config.sync.raw_dir)
        context_dir = Path(config.sync.context_dir)

        all_records: list[dict[str, Any]] = []

        # 同步每个飞书数据源
        for base_config in config.feishu.bases:
            collector = FeishuCollector(
                name=base_config["name"],
                raw_dir=raw_dir / "feishu",
                app_id=config.feishu.app_id,
                app_secret=config.feishu.app_secret,
                app_token=base_config["app_token"],
                table_id=base_config["table_id"],
                field_mapping=base_config.get("fields"),
            )
            records = await collector.collect()
            # 标记为原生内容
            for record in records:
                record["content_type"] = "created"
                record["source"] = "feishu"
            all_records.extend(records)

        if not all_records:
            return "未获取到任何数据。请检查数据源配置。"

        # 路由数据到不同主题
        router = RouterEngine(config)
        routed = router.route_batch(all_records)

        # 写入 Context 文件（使用新的 Obsidian 格式和目录结构）
        written_files = router.write_obsidian_files(
            routed,
            context_dir,
            content_type="created",
        )

        # 生成摘要
        summary_lines = ["## 同步完成 ✓", ""]
        summary_lines.append(f"**总记录数**: {len(all_records)}")
        summary_lines.append("")
        summary_lines.append("**按主题分布**:")
        for topic_key, records in routed.items():
            if records:
                topic = router.topics.get(topic_key)
                topic_name = topic.name if topic else topic_key
                summary_lines.append(f"- {topic_name}: {len(records)} 条")
        summary_lines.append("")
        summary_lines.append(f"**写入文件数**: {len(written_files)}")
        summary_lines.append("")
        summary_lines.append("**存储位置**: `data/context/__created/`")
        summary_lines.append("")
        summary_lines.append("**文件列表**:")
        for f in written_files[:10]:  # 最多显示10个
            summary_lines.append(f"- {f.name}")
        if len(written_files) > 10:
            summary_lines.append(f"- ... 等共 {len(written_files)} 个文件")

        return "\n".join(summary_lines)

    except Exception as e:
        return f"同步失败: {type(e).__name__}: {e}"


@mcp.tool()
async def list_topics() -> str:
    """列出所有可用的 Context 主题及其描述.

    Returns:
        主题列表的 Markdown 格式文本
    """
    try:
        config = get_config()
        topic_manager = get_topic_manager()
        topics = topic_manager.get_all_topics()

        if not topics:
            return "未配置任何主题。"

        lines = ["## 可用主题", ""]

        # 核心主题
        if config.context.core_topics:
            lines.append("### 核心主题")
            lines.append("")
            for key, topic in config.context.core_topics.items():
                lines.append(f"**{topic.name}** (`{key}`)")
                lines.append(f"> {topic.description}")
                if topic.keywords:
                    lines.append(f"关键词: {', '.join(topic.keywords[:5])}...")
                lines.append("")

        # 自定义主题
        if config.context.custom_topics:
            lines.append("### 自定义主题")
            lines.append("")
            for key, topic in config.context.custom_topics.items():
                lines.append(f"**{topic.name}** (`{key}`)")
                lines.append(f"> {topic.description}")
                lines.append("")

        # 动态主题
        if topic_manager.dynamic_topics:
            lines.append("### 动态主题 (AI 自动创建)")
            lines.append("")
            for key, topic in topic_manager.dynamic_topics.items():
                if key not in config.context.get_all_topics():
                    lines.append(f"**{topic.name}** (`{key}`)")
                    lines.append(f"> {topic.description}")
                    lines.append(f"内容数: {topic.content_count}")
                    lines.append("")

        return "\n".join(lines)

    except Exception as e:
        return f"获取主题列表失败: {type(e).__name__}: {e}"


@mcp.tool()
async def get_context(topic: str | None = None) -> str:
    """获取指定主题的 Context 内容，用于回答相关问题.

    Args:
        topic: 主题 key（如 'personal', 'writing', 'tasks'），不传则返回所有主题

    Returns:
        Context 内容的 Markdown 文本
    """
    try:
        config = get_config()
        context_dir = Path(config.sync.context_dir)

        if not context_dir.exists():
            return "Context 目录不存在。请先运行 sync_data 进行同步。"

        # 如果指定了主题，只读取该主题
        if topic:
            # 尝试在新旧两个位置查找
            search_paths = [
                context_dir / "__created" / topic,
                context_dir / "__collected" / topic,
                context_dir / topic,  # 旧位置
            ]

            for topic_dir in search_paths:
                if topic_dir.exists():
                    files = sorted(topic_dir.glob("*.md"), reverse=True)
                    if files:
                        with open(files[0], "r", encoding="utf-8") as f:
                            return f.read()

            return f"主题 '{topic}' 暂无内容。"

        # 未指定主题，返回所有主题的摘要
        lines = ["## Context 总览", ""]

        # 优先读取 __created 目录
        created_dir = context_dir / "__created"
        if created_dir.exists():
            lines.append("### 原生内容 (__created)")
            lines.append("")
            for topic_dir in created_dir.iterdir():
                if topic_dir.is_dir():
                    files = list(topic_dir.glob("*.md"))
                    if files:
                        latest = max(files, key=lambda f: f.stat().st_mtime)
                        with open(latest, "r", encoding="utf-8") as f:
                            content = f.read()
                            preview = "\n".join(content.split("\n")[:5])
                            lines.append(f"**{topic_dir.name}** ({len(files)} 个文件)")
                            lines.append(preview[:200] + "...")
                            lines.append("")

        # 收集内容
        collected_dir = context_dir / "__collected"
        if collected_dir.exists():
            lines.append("### 收集内容 (__collected)")
            lines.append("")
            total_files = 0
            for source_dir in collected_dir.iterdir():
                if source_dir.is_dir():
                    count = len(list(source_dir.rglob("*.md")))
                    total_files += count
                    lines.append(f"- {source_dir.name}: {count} 个文件")
            lines.append(f"**总计**: {total_files} 个文件")
            lines.append("")

        if len(lines) <= 3:
            return "暂无 Context 内容。请先运行 sync_data 或 collect_content。"

        return "\n".join(lines)

    except Exception as e:
        return f"获取 Context 失败: {type(e).__name__}: {e}"


@mcp.tool()
async def query_context(query: str, topics: list[str] | None = None) -> str:
    """跨主题搜索相关内容，返回匹配的记录.

    Args:
        query: 搜索关键词
        topics: 限定搜索的主题列表，不传则搜索所有主题

    Returns:
        匹配的记录列表
    """
    try:
        config = get_config()
        context_dir = Path(config.sync.context_dir)

        if not context_dir.exists():
            return "Context 目录不存在。请先运行 sync_data 进行同步。"

        # 确定要搜索的目录
        search_dirs = []
        if topics:
            for topic in topics:
                search_dirs.extend([
                    context_dir / "__created" / topic,
                    context_dir / "__collected" / topic,
                    context_dir / topic,
                ])
        else:
            # 搜索所有目录
            search_dirs = [
                context_dir / "__created",
                context_dir / "__collected",
                context_dir,
            ]

        matches: list[tuple[str, str, float]] = []  # (topic, line, score)
        query_lower = query.lower()

        for search_dir in search_dirs:
            if not search_dir.exists():
                continue

            # 递归搜索所有 markdown 文件
            for file in search_dir.rglob("*.md"):
                try:
                    with open(file, "r", encoding="utf-8") as f:
                        content = f.read()

                    if query_lower in content.lower():
                        # 提取匹配的段落
                        paragraphs = content.split("\n## ")
                        for para in paragraphs:
                            if query_lower in para.lower():
                                score = para.lower().count(query_lower)
                                topic = file.parent.name
                                matches.append((topic, "## " + para if not para.startswith("## ") else para, score))
                except Exception:
                    continue

        if not matches:
            return f'未找到与 "{query}" 相关的内容。'

        # 按匹配分数排序
        matches.sort(key=lambda x: x[2], reverse=True)

        # 返回前 10 条
        lines = [f"## 搜索: '{query}'", f"**找到 {len(matches)} 条匹配**", ""]

        for topic, content, _ in matches[:10]:
            lines.append(f"### 来自: {topic}")
            lines.append(content[:500] + "..." if len(content) > 500 else content)
            lines.append("")
            lines.append("---")
            lines.append("")

        return "\n".join(lines)

    except Exception as e:
        return f"搜索失败: {type(e).__name__}: {e}"


@mcp.tool()
async def search_recent(days: int = 7, topic: str | None = None) -> str:
    """搜索最近 N 天的内容，返回按时间排序的记录.

    Args:
        days: 搜索最近多少天的内容，默认 7 天
        topic: 限定搜索的主题，不传则搜索所有主题

    Returns:
        匹配的记录列表，按时间倒序排列
    """
    try:
        config = get_config()
        context_dir = Path(config.sync.context_dir)

        if not context_dir.exists():
            return "Context 目录不存在。请先运行 sync_data 进行同步。"

        # 计算时间阈值
        cutoff_date = datetime.now() - timedelta(days=days)

        # 确定要搜索的目录
        search_dirs = []
        if topic:
            search_dirs = [
                context_dir / "__created" / topic,
                context_dir / "__collected" / topic,
                context_dir / topic,
            ]
        else:
            search_dirs = [context_dir / "__created", context_dir / "__collected", context_dir]

        matches: list[tuple[str, str, datetime]] = []  # (topic, content, date)

        for search_dir in search_dirs:
            if not search_dir.exists():
                continue

            for file in search_dir.rglob("*.md"):
                try:
                    # 检查文件修改时间
                    mtime = datetime.fromtimestamp(file.stat().st_mtime)
                    if mtime < cutoff_date:
                        continue

                    with open(file, "r", encoding="utf-8") as f:
                        content = f.read()

                    # 提取日期信息（优先从 frontmatter 解析）
                    file_date = mtime
                    for line in content.split("\n")[:30]:  # 检查前30行
                        if line.startswith("created:"):
                            try:
                                date_str = line.split(":", 1)[1].strip().strip('"')
                                file_date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                                break
                            except (ValueError, IndexError):
                                pass

                    if file_date >= cutoff_date:
                        topic_name = file.parent.name
                        matches.append((topic_name, content, file_date))
                except Exception:
                    continue

        if not matches:
            return f"最近 {days} 天内没有新内容。"

        # 按日期倒序排序
        matches.sort(key=lambda x: x[2], reverse=True)

        # 生成结果
        lines = [f"## 最近 {days} 天的内容", f"**找到 {len(matches)} 条记录**", ""]

        for topic_key, content, date in matches[:20]:  # 最多返回 20 条
            lines.append(f"### {topic_key} ({date.strftime('%Y-%m-%d')})")
            # 只显示前 800 字符
            preview = content[:800] + "..." if len(content) > 800 else content
            lines.append(preview)
            lines.append("")
            lines.append("---")
            lines.append("")

        return "\n".join(lines)

    except Exception as e:
        return f"搜索失败: {type(e).__name__}: {e}"


@mcp.tool()
async def add_note(content: str, topic: str, title: str | None = None, append_to: str | None = None) -> str:
    """快速添加笔记到指定主题.

    支持两种模式：
    - 新建模式（默认）：创建新的笔记文件
    - 追加模式（指定 append_to）：追加内容到已有文件末尾

    Args:
        content: 笔记内容
        topic: 主题 key（如 'reflection', 'writing', 'tasks'）
        title: 笔记标题，不传则使用时间戳
        append_to: 要追加到的已有文件路径（相对于 context_dir 或绝对路径），指定此参数时进入追加模式

    Returns:
        添加结果
    """
    try:
        config = get_config()
        context_dir = Path(config.sync.context_dir).resolve()
        topic_manager = get_topic_manager()

        # 验证主题是否存在，不存在则自动创建
        if not topic_manager.topic_exists(topic):
            # 自动创建主题
            topic_name = topic.replace("-", " ").title()
            topic_manager.create_topic(
                key=topic,
                name=topic_name,
                description=f"{topic_name}相关的笔记和想法",
                auto_created=True,
            )

        now = datetime.now()
        note_title = title or f"笔记 {now.strftime('%H:%M')}"

        # === 追加模式 ===
        if append_to:
            append_path = Path(append_to)
            if not append_path.is_absolute():
                append_path = context_dir / append_to
            
            if not append_path.exists():
                return f"追加失败: 文件不存在 {append_path}"
            
            # 构造追加内容（带分隔线和时间戳）
            append_content = (
                f"\n\n---\n\n"
                f"## 补充记录 ({now.strftime('%Y-%m-%d %H:%M')})\n\n"
                f"{content}\n"
            )
            
            with open(append_path, "a", encoding="utf-8") as f:
                f.write(append_content)
            
            # 更新主题内容计数
            topic_manager.increment_content_count(topic)

            # iWiki 同步信息
            iwiki_topic_mapping = {
                "personal": 4018520218,
                "writing": 4018520218,
                "preferences": 4018520218,
                "reflection": 4018520218,
                "ideas": 4018520218,
                "work": 4018520220,
                "tasks": 4018520220,
                "product": 4018520220,
                "ai": 4018520230,
                "reading": 4018520231,
            }
            iwiki_folder_id = iwiki_topic_mapping.get(topic, 4018520218)

            return (
                f"✅ 笔记已追加到已有文件 '{append_path.name}'\n"
                f"文件: {append_path}\n"
                f"模式: 追加\n"
                f"\n"
                f"📌 iWiki 同步信息:\n"
                f"  - 模式: 追加到已有文档 (使用 saveDocumentParts)\n"
                f"  - spaceid: 4010703137\n"
                f"  - parentid: {iwiki_folder_id}\n"
                f"  - title: {note_title}\n"
                f"  - 请同步调用 iWiki saveDocumentParts 工具追加内容到已有文档"
            )

        # === 新建模式 ===
        # 使用 Obsidian 格式化器
        formatter = ObsidianFormatter()

        # 生成 Obsidian 格式内容
        formatted_content = formatter.format_note(
            title=note_title,
            content=content,
            topic=topic,
            source_type="created",
            source="mcp",
        )

        # 写入 __created 目录
        target_dir = context_dir / "__created" / topic
        target_dir.mkdir(parents=True, exist_ok=True)

        safe_title = "".join(c if c.isalnum() or c in "_-" else "_" for c in note_title)
        filename = f"{topic}_{now.strftime('%Y%m%d')}_{safe_title[:20]}.md"
        filepath = target_dir / filename

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(formatted_content)

        # 更新主题内容计数
        topic_manager.increment_content_count(topic)

        # iWiki 同步信息 - topic 到 iWiki 文件夹 ID 的映射
        iwiki_topic_mapping = {
            "personal": 4018520218,
            "writing": 4018520218,
            "preferences": 4018520218,
            "reflection": 4018520218,
            "ideas": 4018520218,
            "work": 4018520220,
            "tasks": 4018520220,
            "product": 4018520220,
            "ai": 4018520230,
            "reading": 4018520231,
        }
        iwiki_folder_id = iwiki_topic_mapping.get(topic, 4018520218)

        return (
            f"✅ 笔记已添加到 '__created/{topic}'\n"
            f"文件: {filepath}\n"
            f"格式: Obsidian 兼容\n"
            f"\n"
            f"📌 iWiki 同步信息:\n"
            f"  - spaceid: 4010703137\n"
            f"  - parentid: {iwiki_folder_id}\n"
            f"  - title: {note_title}\n"
            f"  - contenttype: MD\n"
            f"  - 请同步调用 iWiki createDocument 工具完成双写"
        )

    except Exception as e:
        return f"添加失败: {type(e).__name__}: {e}"


@mcp.tool()
async def excel_to_csv(file_path: str, output_dir: str | None = None) -> str:
    """将 Excel 文件的所有 sheet 导出为 CSV 文件，返回内容后自动删除临时 CSV 文件.

    适用于需要读取 Excel 内容的场景，会将每个 sheet 转为 CSV 并读取内容，
    最后自动清理生成的临时 CSV 文件。

    Args:
        file_path: Excel 文件的路径（支持 .xlsx / .xls），可以是相对于项目根目录的路径或绝对路径
        output_dir: CSV 输出目录（可选，默认与 Excel 文件同目录）

    Returns:
        所有 sheet 的 CSV 内容，以 Markdown 格式返回
    """
    # 注意：实际处理完全在内存中完成，不会在磁盘上生成临时文件
    try:
        import openpyxl
        import csv
        import io

        excel_path = Path(file_path).resolve()
        if not excel_path.exists():
            return f"❌ 文件不存在: {excel_path}"

        if not excel_path.suffix.lower() in (".xlsx", ".xls"):
            return f"❌ 不支持的文件格式: {excel_path.suffix}，仅支持 .xlsx / .xls"

        # 确定输出目录
        out_dir = Path(output_dir).resolve() if output_dir else excel_path.parent
        out_dir.mkdir(parents=True, exist_ok=True)

        # 打开 Excel 文件（处理样式兼容性问题）
        # 某些 Excel 文件包含 openpyxl 无法解析的样式（如 Fill），需要绕过
        import warnings

        wb = None

        # 方法1: 直接加载 (read_only 模式，不解析样式)
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except (TypeError, ValueError, KeyError, IndexError):
            pass

        # 方法2: Patch stylesheet 解析，跳过有问题的样式 (read_only 模式)
        if wb is None:
            try:
                from openpyxl.reader import excel as excel_reader
                from openpyxl.styles import stylesheet as ss_module

                _orig_apply = getattr(ss_module, "apply_stylesheet", None)
                _orig_apply_in_reader = getattr(excel_reader, "apply_stylesheet", None)

                def _safe_apply_stylesheet(*args, **kwargs):
                    """安全版本的 apply_stylesheet，遇到样式错误时跳过"""
                    try:
                        if _orig_apply:
                            return _orig_apply(*args, **kwargs)
                    except (TypeError, ValueError, KeyError, IndexError):
                        pass

                ss_module.apply_stylesheet = _safe_apply_stylesheet
                if hasattr(excel_reader, "apply_stylesheet"):
                    excel_reader.apply_stylesheet = _safe_apply_stylesheet
                try:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                except (TypeError, ValueError, KeyError, IndexError):
                    # read_only 也失败了，尝试普通模式
                    try:
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            wb = openpyxl.load_workbook(excel_path, data_only=True)
                    except (TypeError, ValueError, KeyError, IndexError):
                        pass
                finally:
                    # 恢复原始函数
                    if _orig_apply:
                        ss_module.apply_stylesheet = _orig_apply
                    if _orig_apply_in_reader and hasattr(excel_reader, "apply_stylesheet"):
                        excel_reader.apply_stylesheet = _orig_apply_in_reader
            except Exception:
                raise

        if wb is None:
            return "❌ 无法打开 Excel 文件，所有加载方式均失败"
        sheet_names = wb.sheetnames

        if not sheet_names:
            wb.close()
            return "❌ Excel 文件中没有任何 sheet"

        results = []

        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            # 收集行数据（纯内存操作，不写磁盘）
            rows_data = []
            try:
                for row in ws.iter_rows(values_only=True):
                    # 将 None 转为空字符串
                    rows_data.append([str(cell) if cell is not None else "" for cell in row])
            except (IndexError, TypeError, KeyError) as row_err:
                # 如果 values_only 模式失败，尝试逐个单元格读取
                rows_data = []
                try:
                    for row in ws.iter_rows():
                        row_values = []
                        for cell in row:
                            try:
                                val = cell.value
                                row_values.append(str(val) if val is not None else "")
                            except (IndexError, TypeError, KeyError):
                                row_values.append("")
                        rows_data.append(row_values)
                except (IndexError, TypeError, KeyError):
                    results.append(f"### 📄 Sheet: {sheet_name}\n\n_(读取失败: {row_err})_\n")
                    continue

            if not rows_data:
                results.append(f"### 📄 Sheet: {sheet_name}\n\n_(空白 sheet)_\n")
                continue

            # 在内存中生成 CSV 内容（不写磁盘，无需清理）
            csv_buffer = io.StringIO()
            writer = csv.writer(csv_buffer)
            writer.writerows(rows_data)
            csv_content = csv_buffer.getvalue()

            # 生成 Markdown 表格预览（前20行）
            preview_rows = rows_data[:20]
            if preview_rows:
                # 表头
                header = preview_rows[0]
                md_table = "| " + " | ".join(header) + " |\n"
                md_table += "| " + " | ".join(["---"] * len(header)) + " |\n"
                # 数据行
                for row in preview_rows[1:]:
                    # 确保列数与表头一致
                    padded_row = row + [""] * (len(header) - len(row)) if len(row) < len(header) else row[:len(header)]
                    md_table += "| " + " | ".join(padded_row) + " |\n"

                total_rows = len(rows_data) - 1  # 减去表头
                truncated_msg = f"\n\n> ⚠️ 仅显示前 19 行数据，共 {total_rows} 行" if total_rows > 19 else ""

                results.append(
                    f"### 📄 Sheet: {sheet_name} ({total_rows} 行数据)\n\n"
                    f"{md_table}{truncated_msg}\n"
                )
            else:
                results.append(f"### 📄 Sheet: {sheet_name}\n\n_(空白 sheet)_\n")

        wb.close()

        # 组装完整结果
        header_msg = (
            f"## 📊 Excel 文件解析完成\n\n"
            f"**源文件**: `{excel_path}`\n"
            f"**Sheet 数量**: {len(sheet_names)}\n"
            f"**Sheet 列表**: {', '.join(sheet_names)}\n\n"
            f"---\n\n"
        )

        full_result = header_msg + "\n---\n\n".join(results)

        return full_result

    except ImportError:
        return (
            "❌ 缺少 openpyxl 依赖，请安装：\n\n"
            "```bash\npip install openpyxl\n```"
        )
    except Exception as e:
        return f"❌ 处理 Excel 文件失败: {type(e).__name__}: {e}"


@mcp.tool()
async def collect_content(
    url: str,
    html: str | None = None,
    topic: str | None = None,
    download_images: bool = True,
) -> str:
    """收集网页内容到知识库（完整内容提取）.

    只需提供 URL，自动完成：
    - 自动抓取网页内容（也可手动传入 HTML）
    - 完整正文提取（保留原始内容，不做摘要）
    - 元数据解析（标题、作者、发布时间）
    - 图片本地化（可选）
    - 保存到知识库

    Args:
        url: 网页 URL
        html: 网页 HTML 内容（可选，不传则自动从 URL 抓取）
        topic: 目标主题（可选，不传则自动推断为 'reading'）
        download_images: 是否下载图片到本地（默认 True）

    Returns:
        处理结果摘要
    """
    try:
        from pa.extractors import WebExtractor, ImageHandler
        from pa.extractors.sites import WechatExtractor, ZhihuExtractor
        from urllib.parse import urlparse

        config = get_config()
        raw_dir = Path(config.sync.raw_dir).resolve()
        context_dir = Path(config.sync.context_dir).resolve()
        topic_manager = get_topic_manager()

        # 0. 如果未提供 HTML，自动从 URL 抓取
        fetch_msg = ""
        if not html:
            try:
                is_reddit = WebExtractor._is_reddit_url(url)
                needs_js = WebExtractor._needs_js_rendering(url)
                html = await WebExtractor.fetch_html(url)
                if is_reddit:
                    fetch_msg = "🔗 Reddit 内容已通过 RSS/JSON API 抓取"
                elif needs_js:
                    fetch_msg = "🎭 HTML 已通过 Playwright（无头浏览器）抓取"
                else:
                    fetch_msg = "🌐 HTML 已自动抓取"
            except RuntimeError as e:
                return f"❌ 无法抓取网页: {e}"

        # 1. 选择合适的提取器
        domain = urlparse(url).netloc.lower()
        
        if "mp.weixin.qq.com" in domain or "weixin.qq.com" in domain:
            extractor = WechatExtractor()
        elif "zhihu.com" in domain:
            extractor = ZhihuExtractor()
        else:
            extractor = WebExtractor()
        
        # 2. 提取完整内容（html 已确保存在）
        extracted = await extractor.extract(url, html)
        
        # 3. 处理图片（如果启用）
        image_infos = []
        processed_content = extracted.content
        
        if download_images and extracted.images:
            assets_dir = context_dir / "__collected" / "assets" / "images"
            image_handler = ImageHandler(assets_dir=assets_dir)
            processed_content, image_infos = await image_handler.process_content(
                extracted.content, source_url=url
            )
        
        # 4. 确定主题
        final_topic = topic or "reading"
        
        if not topic_manager.topic_exists(final_topic):
            topic_name = final_topic.replace("-", " ").title()
            topic_manager.create_topic(
                key=final_topic,
                name=topic_name,
                description=f"{topic_name}相关的收藏文章",
                auto_created=True,
            )
        
        # 5. 生成 Obsidian 格式内容（完整正文）
        now = datetime.now()
        safe_title = "".join(c if c.isalnum() or c in "_-" else "_" for c in extracted.title)[:50]
        
        # 生成知识库内容（Obsidian 格式，包含完整正文）
        obsidian_content = f"""---
created: "{now.strftime('%Y-%m-%d %H:%M:%S')}"
modified: "{now.strftime('%Y-%m-%d %H:%M:%S')}"
tags: ["{final_topic}", "collected", "{extracted.site_type}"]
source-type: collected
source: {extracted.site_type}
url: "{url}"
author: "{extracted.author or ''}"
publish_date: "{extracted.publish_date or ''}"
status: collected
---

# {extracted.title}

> **来源**: {extracted.site_name or extracted.site_type}
> **作者**: {extracted.author or '未知'}
> **发布日期**: {extracted.publish_date or '未知'}
> **原文链接**: [{url}]({url})

---

{processed_content}

---

*收集于 {now.strftime('%Y-%m-%d %H:%M:%S')}*
"""
        
        # 6. 保存原始内容到 raw/collect/
        collect_dir = raw_dir / "collect"
        collect_dir.mkdir(parents=True, exist_ok=True)
        raw_filename = f"{extracted.site_type}_{now.strftime('%Y%m%d_%H%M%S')}_{safe_title}.md"
        raw_filepath = collect_dir / raw_filename
        
        with open(raw_filepath, "w", encoding="utf-8") as f:
            f.write(f"# {extracted.title}\n\n**URL**: {url}\n\n{processed_content}")
        
        # 7. 保存到知识库
        topic_dir = context_dir / "__collected" / extracted.site_type / final_topic
        topic_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{safe_title[:30]}_{now.strftime('%Y%m%d')}.md"
        filepath = topic_dir / filename
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(obsidian_content)
        
        # 8. 更新计数
        topic_manager.increment_content_count(final_topic)
        
        # 返回结果
        image_msg = f"\n🖼️ 图片: 已下载 {len(image_infos)} 张到本地" if image_infos else ""
        fetch_line = f"\n{fetch_msg}" if fetch_msg else ""
        
        return f"""✅ 内容收集完成
{fetch_line}
📄 标题: {extracted.title}
👤 作者: {extracted.author or '未知'}
🏠 来源: {extracted.site_name or extracted.site_type}
📅 发布: {extracted.publish_date or '未知'}
📊 字数: {extracted.word_count}
🏷️ 类型: {extracted.site_type}
📍 主题: {final_topic}{image_msg}

💾 存储位置:
   - 原始数据: {raw_filepath}
   - 知识库: {filepath}

🔧 提取器: {extractor.name}
"""

    except ImportError as e:
        return f"""⚠️ 内容提取功能需要额外依赖

缺少: {e}

请安装依赖:
pip install trafilatura beautifulsoup4
"""
    except Exception as e:
        return f"❌ 收集失败: {type(e).__name__}: {e}"


@mcp.tool()
async def manage_topic(
    action: str,
    topic_key: str | None = None,
    name: str | None = None,
    description: str | None = None,
    keywords: list[str] | None = None,
    parent: str | None = None,
) -> str:
    """管理主题 - 创建、更新、删除或查询主题.

    Args:
        action: 操作类型 (create | update | delete | list | suggest_merge)
        topic_key: 主题 key（create/update/delete 时需要）
        name: 主题显示名称（create/update 时使用）
        description: 主题描述（create/update 时使用）
        keywords: 关键词列表（create/update 时使用）
        parent: 父主题 key（create 时使用，用于层级关系）

    Returns:
        操作结果

    Examples:
        # 创建新主题
        manage_topic(action="create", topic_key="agent-design", name="Agent 设计")

        # 列出所有主题
        manage_topic(action="list")

        # 建议合并的相似主题
        manage_topic(action="suggest_merge")
    """
    try:
        topic_manager = get_topic_manager()

        if action == "list":
            topics = topic_manager.get_all_topics()
            lines = [f"## 主题列表 ({len(topics)} 个)", ""]
            lines.append("| Key | 名称 | 关键词数量 | 类型 |")
            lines.append("|-----|------|-----------|------|")

            config = get_config()
            for key, topic in topics.items():
                topic_type = "核心" if key in config.context.core_topics else "动态"
                lines.append(f"| {key} | {topic.name} | {len(topic.keywords)} | {topic_type} |")

            return "\n".join(lines)

        elif action == "create":
            if not topic_key or not name:
                return "❌ 创建主题需要提供 topic_key 和 name"

            if topic_manager.topic_exists(topic_key):
                return f"❌ 主题 '{topic_key}' 已存在"

            topic = topic_manager.create_topic(
                key=topic_key,
                name=name,
                description=description or f"{name}相关的内容",
                keywords=keywords or [name],
                parent=parent,
                auto_created=False,
            )

            return f"""✅ 主题已创建

Key: {topic_key}
名称: {name}
描述: {topic.description}
关键词: {', '.join(topic.keywords[:5])}
父主题: {parent or '无'}
"""

        elif action == "update":
            if not topic_key:
                return "❌ 更新主题需要提供 topic_key"

            if topic_key not in topic_manager.dynamic_topics:
                return f"❌ 只能更新动态创建的主题（{topic_key} 是核心主题）"

            topic = topic_manager.dynamic_topics[topic_key]
            if name:
                topic.name = name
            if description:
                topic.description = description
            if keywords:
                topic.keywords = keywords
            if parent:
                topic.parent = parent

            topic.updated_at = datetime.now().isoformat()
            topic_manager._save_dynamic_topics()

            return f"✅ 主题 '{topic_key}' 已更新"

        elif action == "delete":
            if not topic_key:
                return "❌ 删除主题需要提供 topic_key"

            if topic_key not in topic_manager.dynamic_topics:
                return f"❌ 只能删除动态创建的主题"

            del topic_manager.dynamic_topics[topic_key]
            topic_manager._save_dynamic_topics()

            return f"✅ 主题 '{topic_key}' 已删除"

        elif action == "suggest_merge":
            suggestions = topic_manager.suggest_merge_topics()

            if not suggestions:
                return "未找到相似度高的主题对，无需合并。"

            lines = ["## 建议合并的主题对", ""]
            lines.append("| 主题1 | 主题2 | 相似度 |")
            lines.append("|-------|-------|--------|")

            for t1, t2, sim in suggestions[:10]:
                lines.append(f"| {t1} | {t2} | {sim:.2%} |")

            lines.append("")
            lines.append("💡 提示: 使用 `manage_topic(action='delete')` 删除重复主题")

            return "\n".join(lines)

        else:
            return f"❌ 未知操作: {action}。支持的操作: create, update, delete, list, suggest_merge"

    except Exception as e:
        return f"操作失败: {type(e).__name__}: {e}"


# ========== Ship-Learn-Next 工具 ==========

@mcp.tool()
async def create_ship_plan(
    topic: str,
    focus: str | None = None,
    duration_weeks: int = 4,
) -> str:
    """基于知识库内容创建 Ship-Learn-Next 实践计划.

    将知识转化为实践的迭代学习计划，遵循 SHIP(交付) → LEARN(反思) → NEXT(迭代) 循环。

    Args:
        topic: 知识来源主题 key（如 'reading', 'ai', 'product'）
        focus: 关注的具体知识点或方向（可选）
        duration_weeks: 计划周期周数（默认4周）

    Returns:
        创建结果摘要，包含计划文件路径

    Example:
        create_ship_plan(topic="reading", focus="Prompt Engineering", duration_weeks=4)
    """
    try:
        config = get_config()
        context_dir = Path(config.sync.context_dir).resolve()
        topic_manager = get_topic_manager()

        # 验证主题
        if not topic_manager.topic_exists(topic):
            available = ", ".join(topic_manager.get_all_topics().keys())
            return f"主题 '{topic}' 不存在。可用主题: {available}"

        # 读取知识内容（优先从新目录读取）
        topic_dirs = [
            context_dir / "__created" / topic,
            context_dir / "__collected" / topic,
            context_dir / topic,
        ]

        knowledge_content = ""
        for topic_dir in topic_dirs:
            if topic_dir.exists():
                files = sorted(topic_dir.glob("*.md"), reverse=True)
                if files:
                    with open(files[0], "r", encoding="utf-8") as f:
                        knowledge_content = f.read()
                        break

        if not knowledge_content:
            return f"主题 '{topic}' 暂无内容。请先运行 sync_data 或 collect_content。"

        # 如果指定了 focus，尝试提取相关内容
        if focus:
            lines = knowledge_content.split("\n")
            relevant_lines = []
            for line in lines:
                if focus.lower() in line.lower():
                    relevant_lines.append(line)
            if relevant_lines:
                knowledge_content = "\n".join(relevant_lines[:20])

        # 导入 planner
        try:
            from planner import ShipLearnPlanner
        except ImportError:
            planner_path = SKILL_DIR / "planner.py"
            if planner_path.exists():
                import importlib.util
                spec = importlib.util.spec_from_file_location("planner", planner_path)
                planner_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(planner_module)
                ShipLearnPlanner = planner_module.ShipLearnPlanner
            else:
                return "错误: 无法加载 planner 模块"

        # 创建计划
        planner = ShipLearnPlanner()
        plan_name = focus or f"{topic}学习实践"

        plan_data = planner.create_plan(
            plan_name=plan_name,
            source_topic=topic,
            knowledge_content=knowledge_content,
            learning_goal=f"将{topic}主题中的知识转化为实际能力" if not focus else f"掌握并应用{focus}",
            duration_weeks=duration_weeks,
        )

        # 生成结果摘要
        lines = [
            f"## ✅ Ship-Learn-Next 计划已创建",
            "",
            f"**计划名称**: {plan_data['plan_name']}",
            f"**知识来源**: {plan_data['source_topic']}",
            f"**计划周期**: {plan_data['duration_weeks']} 周",
            f"**计划文件**: `{plan_data['plan_file']}`",
            "",
            "### 📋 迭代概览",
            "",
        ]

        for week in plan_data['weeks']:
            lines.append(f"**Week {week['number']}** ({week['start_date']} - {week['end_date']})")
            lines.append(f"- SHIP: {week['ship_task'][:60]}...")
            lines.append("")

        lines.extend([
            "### 🚀 下一步",
            "",
            "1. 查看完整计划文件",
            f"2. 开始 Week 1 的 SHIP 任务",
            f"3. 完成后使用 `track_iteration` 记录进度",
        ])

        return "\n".join(lines)

    except Exception as e:
        return f"创建计划失败: {type(e).__name__}: {e}"


@mcp.tool()
async def track_iteration(
    plan_id: str,
    week: int,
    ship_result: str,
    learn_reflection: str,
    next_adjustment: str = "",
) -> str:
    """记录 Ship-Learn-Next 计划的迭代执行结果.

    Args:
        plan_id: 计划ID（如 'Plan-Prompt工程-20260228' 或简写 'Prompt工程'）
        week: 周次（1-4）
        ship_result: SHIP 交付物描述（实际完成了什么）
        learn_reflection: LEARN 反思内容（什么有效/无效/发现/改进）
        next_adjustment: NEXT 调整建议（可选）

    Returns:
        记录结果

    Example:
        track_iteration(
            plan_id="Plan-Prompt工程-20260228",
            week=1,
            ship_result="完成了3个工作场景的Prompt优化，平均效率提升30%",
            learn_reflection="发现简洁的指令效果更好，复杂模板反而容易出错"
        )
    """
    try:
        # 导入 tracker
        try:
            from tracker import IterationTracker
        except ImportError:
            tracker_path = SKILL_DIR / "tracker.py"
            if tracker_path.exists():
                import importlib.util
                spec = importlib.util.spec_from_file_location("tracker", tracker_path)
                tracker_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(tracker_module)
                IterationTracker = tracker_module.IterationTracker
            else:
                return "错误: 无法加载 tracker 模块"

        tracker = IterationTracker()
        result = tracker.track_iteration(
            plan_id=plan_id,
            week=week,
            ship_result=ship_result,
            learn_reflection=learn_reflection,
            next_adjustment=next_adjustment,
        )

        if "error" in result:
            return f"记录失败: {result['error']}"

        # 获取最新状态
        status = tracker.get_plan_status(plan_id)

        lines = [
            f"## ✅ Week {week} 迭代已记录",
            "",
            f"**计划**: {plan_id}",
            f"**总体进度**: {status.get('progress', 'N/A')}",
            f"**状态**: {'已完成' if status.get('status') == 'completed' else '进行中'}",
            "",
            "### 📝 记录内容",
            "",
            f"**SHIP**: {ship_result[:100]}...",
            "",
            f"**LEARN**: {learn_reflection[:100]}...",
            "",
            "### 🚀 下一步",
            "",
        ]

        if week < 4:
            lines.append(f"继续 Week {week + 1} 的 SHIP 任务！")
        else:
            lines.append("恭喜完成整个计划！可以使用 `export_plan_to_ppt` 导出分享。")

        return "\n".join(lines)

    except Exception as e:
        return f"记录失败: {type(e).__name__}: {e}"


@mcp.tool()
async def export_plan_to_ppt(
    plan_id: str,
    style: str = "snoopy",
) -> str:
    """将 Ship-Learn-Next 计划导出为 PPT 分享文档.

    Args:
        plan_id: 计划ID（如 'Plan-Prompt工程-20260228'）
        style: 视觉风格（可选: snoopy, manga, oatmeal, cyberpunk 等）

    Returns:
        导出结果，包含 PPT 文件路径

    Example:
        export_plan_to_ppt(plan_id="Plan-Prompt工程-20260228", style="snoopy")
    """
    try:
        # 导入 exporter
        try:
            from exporter import PlanExporter
        except ImportError:
            exporter_path = SKILL_DIR / "exporter.py"
            if exporter_path.exists():
                import importlib.util
                spec = importlib.util.spec_from_file_location("exporter", exporter_path)
                exporter_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(exporter_module)
                PlanExporter = exporter_module.PlanExporter
            else:
                return "错误: 无法加载 exporter 模块"

        exporter = PlanExporter()
        result = exporter.export_to_ppt(
            plan_id=plan_id,
            style=style,
        )

        if "error" in result:
            return f"导出失败: {result['error']}"

        if result.get("success"):
            lines = [
                f"## ✅ PPT 导出成功！",
                "",
                f"**计划**: {plan_id}",
                f"**幻灯片数量**: {result['slide_count']} 页",
                f"**风格**: {result['style']}",
                f"**输出文件**: `{result['output_file']}`",
                "",
                "### 📊 PPT 包含内容",
                "",
                "1. 封面 - 计划名称和概述",
                "2. 学习目标",
                "3. 知识来源摘要",
                "4-7. Week 1-4 的迭代计划",
                "8. 行动号召",
                "",
                "可以直接用于分享和展示！",
            ]
            return "\n".join(lines)
        else:
            return "导出失败: 未知错误"

    except Exception as e:
        return f"导出失败: {type(e).__name__}: {e}"


@mcp.tool()
async def list_ship_plans() -> str:
    """列出所有 Ship-Learn-Next 实践计划.

    Returns:
        计划列表
    """
    try:
        # 导入 tracker
        try:
            from tracker import IterationTracker
        except ImportError:
            tracker_path = SKILL_DIR / "tracker.py"
            if tracker_path.exists():
                import importlib.util
                spec = importlib.util.spec_from_file_location("tracker", tracker_path)
                tracker_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(tracker_module)
                IterationTracker = tracker_module.IterationTracker
            else:
                return "错误: 无法加载 tracker 模块"

        tracker = IterationTracker()
        plans = tracker.list_active_plans()

        if not plans:
            return "暂无 Ship-Learn-Next 计划。使用 `create_ship_plan` 创建第一个计划！"

        lines = [
            f"## 📋 Ship-Learn-Next 计划列表 ({len(plans)} 个)",
            "",
            "| 计划名称 | 创建日期 | 进度 | 状态 |",
            "|---------|---------|------|------|",
        ]

        for plan in plans:
            status_emoji = "✅" if plan['status'] == 'completed' else "🔄"
            lines.append(f"| {plan['name']} | {plan['date']} | {plan['progress']} | {status_emoji} |")

        lines.extend([
            "",
            "### 💡 提示",
            "",
            "- 使用 `create_ship_plan` 创建新计划",
            "- 使用 `track_iteration` 记录进度",
            "- 使用 `export_plan_to_ppt` 导出分享",
        ])

        return "\n".join(lines)

    except Exception as e:
        return f"获取计划列表失败: {type(e).__name__}: {e}"


# ========== 每日资讯采集工具 ==========

# 全局调度器实例缓存（用于交互式多轮对话）
_scheduler_instance = None


def _get_scheduler():
    """获取调度器实例（带缓存，支持交互式多轮筛选）."""
    global _scheduler_instance
    if _scheduler_instance is None:
        from pa.scheduler import DailyDigestScheduler
        config = get_config()
        _scheduler_instance = DailyDigestScheduler(config)
    return _scheduler_instance


def _reset_scheduler(force_new: bool = False):
    """重置调度器实例.

    Args:
        force_new: 如果为 True，同时清除本地缓存（强制全新开始）
    """
    global _scheduler_instance
    if force_new and _scheduler_instance is not None:
        _scheduler_instance.reset_cache()
    _scheduler_instance = None


def _format_article_list(articles: list[dict[str, Any]]) -> str:
    """将文章列表格式化为用户可读的编号列表."""
    if not articles:
        return "暂无新文章。"

    lines: list[str] = []
    for i, article in enumerate(articles, 1):
        title = article.get("title", "无标题")
        source = article.get("source_name", "未知来源")
        url = article.get("url", "")
        desc = article.get("description", "")
        topic = article.get("topic", "")
        pub_date = article.get("pub_date", "")

        lines.append(f"### {i}. {title}")
        lines.append(f"> 📌 来源: {source} | 🏷️ 主题: {topic} | 📅 {pub_date}")
        if desc:
            lines.append(f"\n{desc[:150]}{'...' if len(desc) > 150 else ''}")
        if url:
            lines.append(f"\n🔗 [查看原文]({url})")
        lines.append("")

    return "\n".join(lines)


@mcp.tool()
async def fetch_daily_digest(force_refresh: bool = False) -> str:
    """抓取 RSS 订阅源的最新文章，展示列表供用户筛选.

    采用分批抓取策略避免超时：每次调用抓取一批（4个源），自动缓存结果。
    - 如果还有未抓取的批次，会提示继续调用本工具完成剩余批次
    - 所有批次完成后展示完整文章列表
    - 同一天内重复调用会复用缓存，除非 force_refresh=True

    Args:
        force_refresh: 是否强制刷新（清除今日缓存，重新开始抓取）

    Returns:
        文章列表或分批进度提示
    """
    try:
        scheduler = _get_scheduler()

        # 如果强制刷新，清除缓存重新开始
        if force_refresh:
            _reset_scheduler(force_new=True)
            scheduler = _get_scheduler()

        result = await scheduler.fetch_articles()

        batch_index = result["batch_index"]
        total_batches = result["total_batches"]
        completed = result["completed"]
        batch_articles = result["articles"]
        all_articles = result["all_articles"]
        errors = result["errors"]
        all_errors = result["all_errors"]

        # 如果还有未完成的批次，返回进度提示并提示继续调用
        if not completed:
            lines = [
                "## 📰 资讯抓取进行中...",
                "",
                f"**进度**: 第 {batch_index + 1}/{total_batches} 批完成"
                f"（本批抓到 {len(batch_articles)} 篇，累计 {len(all_articles)} 篇）",
                "",
            ]

            if errors:
                lines.append("**本批次错误**:")
                for err in errors:
                    lines.append(f"- {err}")
                lines.append("")

            remaining = total_batches - batch_index - 1
            lines.extend([
                f"还有 **{remaining}** 批待抓取，请继续调用 `fetch_daily_digest` 完成剩余批次。",
                "",
                "💡 无需传参，工具会自动从上次中断处继续。",
            ])

            return "\n".join(lines)

        # 所有批次完成，展示完整列表
        lines = [
            "## 📰 今日资讯推送",
            "",
            f"共检查 **{result['total_sources']}** 个订阅源（{total_batches} 批次），"
            f"发现 **{len(all_articles)}** 篇新文章：",
            "",
        ]

        if all_articles:
            lines.append(_format_article_list(all_articles))
            lines.extend([
                "---",
                "",
                "### 📋 操作指引",
                "",
                "请告诉我你想加入知识库的文章编号，例如：",
                "- `保存 1, 3, 5` - 保存指定文章",
                "- `全部保存` - 保存所有文章",
                "- `跳过` - 全部跳过不保存",
                "- `继续抓取` - 觉得不够？我再抓一批",
                "",
                "💡 所有展示过的文章（无论是否保存）下次不会重复推送。",
            ])
        else:
            lines.append("暂时没有新的文章。所有已推送的文章都已标记，等待新内容产生。")

        if all_errors:
            lines.extend([
                "",
                "### ⚠️ 部分源抓取失败",
            ])
            for err in all_errors:
                lines.append(f"- {err}")

        return "\n".join(lines)

    except Exception as e:
        return f"❌ 资讯抓取失败: {type(e).__name__}: {e}"


@mcp.tool()
async def save_selected_articles(
    indices: str,
) -> str:
    """保存用户选中的文章到知识库.

    Args:
        indices: 用户选择的文章编号，支持格式：
                - "1,3,5" 或 "1, 3, 5" - 指定编号
                - "all" 或 "全部" - 保存全部
                - "skip" 或 "跳过" - 全部跳过

    Returns:
        保存结果
    """
    try:
        scheduler = _get_scheduler()

        # 解析用户输入
        indices_str = indices.strip().lower()

        if indices_str in ("skip", "跳过", "none", "无"):
            count = scheduler.skip_all()
            return f"✅ 已跳过全部 {count} 篇文章，下次不会重复推送。\n\n💡 如果还想继续看更多文章，可以调用 `fetch_more_articles`。"

        if indices_str in ("all", "全部", "全部保存"):
            # 保存所有文章
            total = len(scheduler._current_batch)
            selected = list(range(1, total + 1))
        else:
            # 解析编号
            try:
                selected = [
                    int(x.strip())
                    for x in indices_str.replace("，", ",").split(",")
                    if x.strip().isdigit()
                ]
            except ValueError:
                return "❌ 无法解析文章编号，请使用逗号分隔的数字，如 `1, 3, 5`"

        if not selected:
            return "❌ 未选择任何文章，请提供文章编号。"

        result = scheduler.save_selected(selected)

        lines = [
            "## ✅ 保存完成",
            "",
            f"**已保存**: {result['saved_count']} 篇文章",
            f"**已标记已读**: {result['marked_seen']} 篇（下次不再推送）",
            "",
        ]

        if result["files_written"]:
            lines.append("### 📁 保存位置")
            for f in result["files_written"]:
                filepath = Path(f)
                lines.append(f"- `{filepath.parent.name}/{filepath.name}`")
            lines.append("")

        lines.extend([
            "---",
            "",
            "💡 接下来你可以：",
            "- `继续抓取` - 还想看更多文章",
            "- 确认完成 - 本次采集结束",
        ])

        return "\n".join(lines)

    except Exception as e:
        return f"❌ 保存失败: {type(e).__name__}: {e}"


@mcp.tool()
async def fetch_more_articles() -> str:
    """继续抓取更多文章（当用户觉得之前的推送数量不够时调用）.

    从上次的位置继续往下抓取每个 RSS 源的更多内容。

    Returns:
        新一批文章列表
    """
    try:
        scheduler = _get_scheduler()
        result = await scheduler.fetch_more()

        articles = result["articles"]
        errors = result["errors"]

        lines = [
            "## 📰 更多资讯",
            "",
            f"继续抓取，又发现 **{result['total_found']}** 篇新文章：",
            "",
        ]

        if articles:
            lines.append(_format_article_list(articles))
            lines.extend([
                "---",
                "",
                "### 📋 操作指引",
                "",
                "请告诉我你想保存的文章编号：",
                "- `保存 1, 3` - 保存指定文章",
                "- `全部保存` - 全部保存",
                "- `跳过` - 全部跳过",
                "- `继续抓取` - 还想看更多",
            ])
        else:
            lines.extend([
                "暂时没有更多新文章了，所有源的内容都已抓取完毕。",
                "",
                "💡 你可以：",
                "- 在 `config.yaml` 中添加更多 RSS 源",
                "- 等待订阅源更新后再次抓取",
            ])

        if errors:
            lines.extend([
                "",
                "### ⚠️ 错误",
            ])
            for err in errors:
                lines.append(f"- {err}")

        return "\n".join(lines)

    except Exception as e:
        return f"❌ 继续抓取失败: {type(e).__name__}: {e}"


@mcp.tool()
async def list_topic_notes(topic: str) -> str:
    """列出指定主题下的所有本地笔记文件.

    用于在添加笔记前查看已有文件，判断是否应追加到已有文件而非创建新文件。

    Args:
        topic: 主题 key（如 'work', 'personal', 'ai'）

    Returns:
        该主题下的文件列表及标题摘要
    """
    try:
        config = get_config()
        context_dir = Path(config.sync.context_dir).resolve()
        target_dir = context_dir / "__created" / topic

        if not target_dir.exists():
            return f"主题 '{topic}' 目录不存在，暂无笔记文件。"

        files = sorted(target_dir.glob("*.md"), key=lambda f: f.stat().st_mtime, reverse=True)
        if not files:
            return f"主题 '{topic}' 目录为空，暂无笔记文件。"

        lines = [
            f"## 📂 主题 '{topic}' 下共有 {len(files)} 个笔记文件",
            "",
        ]

        for f in files:
            # 读取文件前几行提取标题
            try:
                with open(f, "r", encoding="utf-8") as fh:
                    head_lines = []
                    for i, line in enumerate(fh):
                        if i >= 10:
                            break
                        head_lines.append(line.strip())

                # 尝试从 YAML front matter 或 # 标题中提取
                doc_title = f.stem
                for line in head_lines:
                    if line.startswith("# "):
                        doc_title = line[2:].strip()
                        break
                    if line.startswith("title:"):
                        doc_title = line[6:].strip().strip('"').strip("'")
                        break

                # 获取文件大小和修改时间
                stat = f.stat()
                mod_time = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")

                lines.append(f"- **{doc_title}**")
                lines.append(f"  文件: `{f.name}` | 路径: `{f}` | 修改: {mod_time}")
                lines.append("")
            except Exception:
                lines.append(f"- {f.name} (读取失败)")
                lines.append("")

        return "\n".join(lines)

    except Exception as e:
        return f"列出失败: {type(e).__name__}: {e}"


def main() -> None:
    """启动 MCP Server."""
    # 设置正确的工作目录为项目根目录
    import os
    project_root = Path(__file__).parent.parent.parent
    os.chdir(project_root)

    # 注册浏览器自动化工具（Playwright MCP 工具集）
    try:
        from pa.browser_tools import register_browser_tools
        register_browser_tools(mcp)
    except ImportError:
        pass  # playwright 未安装时跳过

    mcp.run()


if __name__ == "__main__":
    main()
