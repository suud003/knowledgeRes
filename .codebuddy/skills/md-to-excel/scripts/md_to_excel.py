#!/usr/bin/env python3
"""Markdown 文档适配到 Excel 的转换脚本.

支持功能：
1. 提取 Markdown 中的所有表格，写入 Excel（每个表格一个 sheet）
2. 将整个 Markdown 文档结构化写入 Excel（标题/内容/层级）
3. 支持自定义列宽、样式、合并单元格等
4. 支持多个 Markdown 文件批量转换

依赖：openpyxl（项目已有）

用法：
    python md_to_excel.py <input.md> [--output output.xlsx] [--mode tables|document|structured]
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────
# Markdown 解析器
# ──────────────────────────────────────────────

def parse_md_tables(md_content: str) -> list[dict[str, Any]]:
    """从 Markdown 内容中提取所有表格.

    Returns:
        [{"title": "表格上方最近的标题", "headers": [...], "rows": [[...], ...]}]
    """
    lines = md_content.split("\n")
    tables: list[dict[str, Any]] = []
    current_title = "Sheet1"
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # 记录最近的标题
        if line.startswith("#"):
            current_title = line.lstrip("#").strip()

        # 检测表格：当前行包含 | 且下一行是分隔行 (|---|...)
        if "|" in line and i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            if re.match(r"^\|[\s\-:|]+\|$", next_line):
                # 解析表头
                headers = _parse_table_row(line)
                i += 2  # 跳过分隔行

                # 解析数据行
                rows = []
                while i < len(lines) and "|" in lines[i].strip():
                    row_line = lines[i].strip()
                    if re.match(r"^\|[\s\-:|]+\|$", row_line):
                        i += 1
                        continue
                    row = _parse_table_row(row_line)
                    # 补齐列数
                    while len(row) < len(headers):
                        row.append("")
                    rows.append(row[: len(headers)])
                    i += 1

                tables.append({
                    "title": _sanitize_sheet_name(current_title),
                    "headers": headers,
                    "rows": rows,
                })
                continue
        i += 1

    return tables


def parse_md_structure(md_content: str) -> list[dict[str, Any]]:
    """将 Markdown 解析为结构化块列表.

    Returns:
        [{"type": "heading|paragraph|list|table|code", "level": int, "content": str, ...}]
    """
    lines = md_content.split("\n")
    blocks: list[dict[str, Any]] = []
    i = 0

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # 空行跳过
        if not stripped:
            i += 1
            continue

        # 标题
        if stripped.startswith("#"):
            match = re.match(r"^(#{1,6})\s+(.*)", stripped)
            if match:
                level = len(match.group(1))
                blocks.append({
                    "type": "heading",
                    "level": level,
                    "content": _strip_md_formatting(match.group(2).strip()),
                })
                i += 1
                continue

        # 表格
        if "|" in stripped and i + 1 < len(lines):
            next_line = lines[i + 1].strip() if i + 1 < len(lines) else ""
            if re.match(r"^\|[\s\-:|]+\|$", next_line):
                headers = _parse_table_row(stripped)
                i += 2
                rows = []
                while i < len(lines) and "|" in lines[i].strip():
                    row_line = lines[i].strip()
                    if re.match(r"^\|[\s\-:|]+\|$", row_line):
                        i += 1
                        continue
                    row = _parse_table_row(row_line)
                    while len(row) < len(headers):
                        row.append("")
                    rows.append(row[: len(headers)])
                    i += 1
                blocks.append({
                    "type": "table",
                    "headers": headers,
                    "rows": rows,
                })
                continue

        # 代码块
        if stripped.startswith("```"):
            lang = stripped[3:].strip()
            code_lines = []
            i += 1
            while i < len(lines) and not lines[i].strip().startswith("```"):
                code_lines.append(lines[i])
                i += 1
            i += 1  # 跳过结束的 ```
            blocks.append({
                "type": "code",
                "language": lang,
                "content": "\n".join(code_lines),
            })
            continue

        # 列表项
        if re.match(r"^[\-\*\+]\s+", stripped) or re.match(r"^\d+\.\s+", stripped):
            list_items = []
            while i < len(lines):
                item_line = lines[i].strip()
                if not item_line:
                    break
                list_match = re.match(r"^(?:[\-\*\+]|\d+\.)\s+(.*)", item_line)
                if list_match:
                    list_items.append(list_match.group(1))
                elif item_line.startswith("  ") or item_line.startswith("\t"):
                    # 续行内容追加到上一项
                    if list_items:
                        list_items[-1] += " " + item_line.strip()
                else:
                    break
                i += 1
            blocks.append({
                "type": "list",
                "items": list_items,
            })
            continue

        # 普通段落 — 每行独立保留，不合并为一行
        para_lines = [stripped]
        i += 1
        while i < len(lines) and lines[i].strip() and not lines[i].strip().startswith("#"):
            if "|" in lines[i].strip() or lines[i].strip().startswith("```"):
                break
            if re.match(r"^(?:[\-\*\+]|\d+\.)\s+", lines[i].strip()):
                break
            para_lines.append(lines[i].strip())
            i += 1
        blocks.append({
            "type": "paragraph",
            "lines": para_lines,  # 保留为多行列表
            "content": " ".join(para_lines),  # 兼容旧接口
        })

    return blocks


# ──────────────────────────────────────────────
# Excel 写入器
# ──────────────────────────────────────────────

# 样式常量
HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
HEADING_FONTS = {
    1: Font(name="微软雅黑", size=16, bold=True, color="FF000000"),
    2: Font(name="微软雅黑", size=14, bold=True, color="FF000000"),
    3: Font(name="微软雅黑", size=12, bold=True, color="FF000000"),
    4: Font(name="微软雅黑", size=11, bold=True, color="FF000000"),
    5: Font(name="微软雅黑", size=11, bold=True, color="FF000000"),
    6: Font(name="微软雅黑", size=11, bold=True, color="FF000000"),
}
# 标题填充：灰色底色
HEADING_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
# 内容填充：浅绿色底色 #EAFAF1
CONTENT_FILL = PatternFill(start_color="FFEAFAF1", end_color="FFEAFAF1", fill_type="solid")
BODY_FONT = Font(name="微软雅黑", size=10, color="FF000000")
CODE_FONT = Font(name="Consolas", size=9, color="FF000000")
CODE_FILL = PatternFill(start_color="FFEAFAF1", end_color="FFEAFAF1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=False, vertical="top")
# 固定的内容区域背景色列数
FIXED_CONTENT_COLS = 8


def write_tables_to_excel(
    tables: list[dict[str, Any]],
    output_path: str | Path,
    *,
    auto_width: bool = True,
) -> Path:
    """将提取的表格列表写入 Excel 文件（每个表格一个 sheet）.

    Args:
        tables: parse_md_tables() 返回的表格列表
        output_path: 输出文件路径
        auto_width: 是否自动调整列宽

    Returns:
        输出文件的 Path 对象
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除默认 sheet

    used_names: set[str] = set()

    for idx, table in enumerate(tables):
        # 生成唯一的 sheet 名
        name = table.get("title", f"Sheet{idx + 1}")
        name = _unique_sheet_name(name, used_names)
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        headers = table["headers"]
        rows = table["rows"]

        # 写表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # 写数据行
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.alignment = WRAP_ALIGNMENT
                cell.border = THIN_BORDER

        # 自动列宽
        if auto_width:
            _auto_adjust_column_width(ws, headers, rows)

        # 冻结首行
        ws.freeze_panes = "A2"
        # 添加自动筛选
        if headers:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    out = Path(output_path)
    wb.save(str(out))
    wb.close()
    return out


def write_document_to_excel(
    blocks: list[dict[str, Any]],
    output_path: str | Path,
    *,
    sheet_name: str = "文档内容",
) -> Path:
    """将结构化文档块写入 Excel（单 sheet，层级缩进布局）.

    布局规则：
    - 不显示"类型""层级"等属性列
    - 标题根据层级占据不同的列位置，同层级标题在同一列
    - 标题下的内容往后偏移一列，体现层级从属关系
    - 多行文字每行占 Excel 一行
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(sheet_name)

    current_row = 1
    # 跟踪当前所处的标题层级，用于决定内容列偏移
    current_heading_level = 0
    # 跟踪上一个块的类型，用于在标题与内容之间插入空行和过渡行
    prev_block_type = None
    # 标记当前是否处于内容区域（已写入过渡行开始），用于在内容结束时关闭过渡
    in_content_section = False
    # 固定内容区域的背景色为 FIXED_CONTENT_COLS 列
    content_max_col = FIXED_CONTENT_COLS

    def _write_transition_row(row_num: int, col: int) -> int:
        """写入一行带浅绿色背景的空过渡行，返回下一行行号.

        过渡行的列数与内容区域保持一致（由 content_max_col 决定），
        确保视觉上过渡行和内容等宽。不使用边框。
        """
        end_col = col + FIXED_CONTENT_COLS - 1
        for c in range(col, end_col + 1):
            cell = ws.cell(row=row_num, column=c, value="")
            cell.fill = CONTENT_FILL
        # 不设置特殊行高，让过渡行和正文行保持一样的默认高度
        return row_num + 1

    def _close_content_section() -> None:
        """关闭内容区域：写入下方过渡行 + 空行."""
        nonlocal current_row, in_content_section, content_max_col
        if in_content_section:
            col = current_heading_level + 1
            # 内容下方：浅绿色过渡行
            current_row = _write_transition_row(current_row, col)
            # 再空一行
            current_row += 1
            in_content_section = False

    def _open_content_section() -> None:
        """开启内容区域：如果尚未开启，写入上方过渡行."""
        nonlocal current_row, in_content_section
        if not in_content_section:
            col = current_heading_level + 1
            # 内容上方：浅绿色过渡行
            current_row = _write_transition_row(current_row, col)
            in_content_section = True

    def _fill_content_row(row_num: int, col: int):
        """为内容行填充背景色（不含边框），固定填充 FIXED_CONTENT_COLS 列."""
        end_col = col + FIXED_CONTENT_COLS - 1
        for c in range(col, end_col + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = CONTENT_FILL

    # 预扫描：为每个标题区域计算其下所有内容块的最大列数
    # key = 块索引, value = 该标题下内容区域所需的最大列数
    _section_max_cols: dict[int, int] = {}
    _scan_heading_idx = -1
    _scan_max = 1
    for _bi, _blk in enumerate(blocks):
        if _blk["type"] == "heading":
            # 保存前一个标题区域的最大列数
            if _scan_heading_idx >= 0:
                _section_max_cols[_scan_heading_idx] = _scan_max
            _scan_heading_idx = _bi
            _scan_max = 1
        elif _blk["type"] == "table":
            _scan_max = max(_scan_max, len(_blk.get("headers", [])))
    if _scan_heading_idx >= 0:
        _section_max_cols[_scan_heading_idx] = _scan_max

    for block_idx, block in enumerate(blocks):
        block_type = block["type"]

        if block_type == "heading":
            # 标题前：如果前面有未关闭的内容区域，先关闭（加过渡行+空行）
            _close_content_section()

            # 连续标题之间不需要额外空行（上一个标题后已有空行）
            # 但如果前一个也是标题，当前 current_row 已经在空行之后了

            level = block.get("level", 1)
            current_heading_level = level
            # 标题写入第 level 列（H1→A列, H2→B列, H3→C列...）
            col = level
            cell = ws.cell(row=current_row, column=col, value=_strip_md_formatting(block["content"]))
            font = HEADING_FONTS.get(level, HEADING_FONTS[6])
            cell.font = font
            cell.fill = HEADING_FILL  # 灰色底色
            cell.border = THIN_BORDER
            # 标题行固定填充10列灰色底色
            FIXED_HEADING_COLS = 10
            heading_end_col = col + FIXED_HEADING_COLS
            for hcol in range(col + 1, heading_end_col):
                hc = ws.cell(row=current_row, column=hcol, value="")
                hc.fill = HEADING_FILL
                hc.border = THIN_BORDER
            current_row += 1
            # 标题后：空一行（作为标题与内容过渡行之间的间距）
            current_row += 1
            prev_block_type = "heading"

        elif block_type == "paragraph":
            col = current_heading_level + 1
            # 开启内容区域（写入上方过渡行）
            _open_content_section()
            lines = block.get("lines", [block.get("content", "")])
            for line in lines:
                if line.strip():
                    cell = ws.cell(row=current_row, column=col, value=_strip_md_formatting(line))
                    cell.font = BODY_FONT
                    cell.alignment = WRAP_ALIGNMENT
                    # 填充整行背景色（无边框）
                    _fill_content_row(current_row, col)
                    current_row += 1
            prev_block_type = "paragraph"

        elif block_type == "list":
            col = current_heading_level + 1
            # 开启内容区域（写入上方过渡行）
            _open_content_section()
            for item in block.get("items", []):
                cell = ws.cell(row=current_row, column=col, value=f"• {_strip_md_formatting(item)}")
                cell.font = BODY_FONT
                cell.alignment = WRAP_ALIGNMENT
                # 填充整行背景色（无边框）
                _fill_content_row(current_row, col)
                current_row += 1
            prev_block_type = "list"

        elif block_type == "table":
            headers = block.get("headers", [])
            rows = block.get("rows", [])
            start_col = current_heading_level + 1
            # 开启内容区域（写入上方过渡行）
            _open_content_section()

            # 写表头（表格保留边框）
            for col_idx, h in enumerate(headers):
                c = ws.cell(row=current_row, column=start_col + col_idx, value=h)
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
                c.border = THIN_BORDER
            # 表头行：表格列之后的空列填充背景色（无边框），对齐固定列数
            table_end_col = start_col + len(headers)
            bg_end_col = start_col + FIXED_CONTENT_COLS
            for extra_col in range(table_end_col, bg_end_col):
                ec = ws.cell(row=current_row, column=extra_col, value="")
                ec.fill = CONTENT_FILL
            current_row += 1

            # 写数据行（表格保留边框）
            for row in rows:
                for col_idx, val in enumerate(row):
                    c = ws.cell(row=current_row, column=start_col + col_idx, value=val)
                    c.font = BODY_FONT
                    c.fill = CONTENT_FILL  # 浅绿色底色
                    c.alignment = WRAP_ALIGNMENT
                    c.border = THIN_BORDER
                # 数据行：表格列之后的空列填充背景色（无边框），对齐固定列数
                for extra_col in range(table_end_col, bg_end_col):
                    ec = ws.cell(row=current_row, column=extra_col, value="")
                    ec.fill = CONTENT_FILL
                current_row += 1
            prev_block_type = "table"

        elif block_type == "code":
            col = current_heading_level + 1
            # 开启内容区域（写入上方过渡行）
            _open_content_section()
            code_content = block.get("content", "")
            # 代码块每行独立写入（无边框）
            for code_line in code_content.split("\n"):
                cell = ws.cell(row=current_row, column=col, value=code_line)
                cell.font = CODE_FONT
                cell.alignment = WRAP_ALIGNMENT
                # 填充整行背景色（无边框）
                _fill_content_row(current_row, col)
                current_row += 1
            prev_block_type = "code"

    # 文档末尾：关闭最后一个内容区域
    _close_content_section()

    # 自动列宽 — 根据实际使用的最大列数设置
    max_col = ws.max_column or 1
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        # 遍历所有有值的单元格计算最大宽度
        max_len = 8
        for row_idx in range(1, current_row):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
        adjusted_width = min(max_len * 1.2 + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 12)

    out = Path(output_path)
    wb.save(str(out))
    wb.close()
    return out


def write_structured_to_excel(
    md_content: str,
    output_path: str | Path,
    *,
    mode: str = "auto",
) -> Path:
    """智能模式：自动判断最佳转换策略.

    Args:
        md_content: Markdown 原文
        output_path: 输出路径
        mode: 转换模式
            - "tables": 仅提取表格
            - "document": 全文档结构化
            - "auto": 自动判断（有表格用 tables，否则用 document）

    Returns:
        输出文件路径
    """
    if mode == "tables":
        tables = parse_md_tables(md_content)
        if not tables:
            raise ValueError("Markdown 中未找到任何表格")
        return write_tables_to_excel(tables, output_path)

    elif mode == "document":
        blocks = parse_md_structure(md_content)
        if not blocks:
            raise ValueError("Markdown 内容为空")
        return write_document_to_excel(blocks, output_path)

    else:  # auto
        tables = parse_md_tables(md_content)
        if tables:
            return write_tables_to_excel(tables, output_path)
        else:
            blocks = parse_md_structure(md_content)
            if not blocks:
                raise ValueError("Markdown 内容为空")
            return write_document_to_excel(blocks, output_path)


# ──────────────────────────────────────────────
# 辅助函数
# ──────────────────────────────────────────────

def _strip_md_formatting(text: str) -> str:
    """清除文本中的 Markdown 格式标记，保留纯文本内容.

    清除的格式包括：
    - **粗体** / __粗体__
    - *斜体* / _斜体_
    - ~~删除线~~
    - `行内代码`
    - [链接文字](url) → 保留链接文字
    - ![图片](url) → 保留图片描述
    - > 引用标记
    - HTML 标签（如 <br>、<b>text</b> 等）
    """
    if not text or not isinstance(text, str):
        return text or ""
    s = text
    # 图片 ![alt](url) → alt
    s = re.sub(r'!\[([^\]]*)\]\([^)]*\)', r'\1', s)
    # 链接 [text](url) → text
    s = re.sub(r'\[([^\]]*)\]\([^)]*\)', r'\1', s)
    # 行内代码 `code` → code
    s = re.sub(r'`([^`]*)`', r'\1', s)
    # 粗斜体 ***text*** 或 ___text___
    s = re.sub(r'\*{3}(.+?)\*{3}', r'\1', s)
    s = re.sub(r'_{3}(.+?)_{3}', r'\1', s)
    # 粗体 **text** 或 __text__
    s = re.sub(r'\*{2}(.+?)\*{2}', r'\1', s)
    s = re.sub(r'_{2}(.+?)_{2}', r'\1', s)
    # 斜体 *text* 或 _text_（注意不要误伤下划线连接词）
    s = re.sub(r'(?<!\w)\*(.+?)\*(?!\w)', r'\1', s)
    s = re.sub(r'(?<!\w)_(.+?)_(?!\w)', r'\1', s)
    # 删除线 ~~text~~
    s = re.sub(r'~~(.+?)~~', r'\1', s)
    # 引用标记 > (行首)
    s = re.sub(r'^>\s?', '', s, flags=re.MULTILINE)
    # HTML 标签
    s = re.sub(r'<[^>]+>', '', s)
    return s.strip()


def _parse_table_row(line: str) -> list[str]:
    """解析 Markdown 表格行，返回单元格列表."""
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    return [_strip_md_formatting(cell.strip()) for cell in line.split("|")]


def _sanitize_sheet_name(name: str) -> str:
    """清理 sheet 名称（Excel 不允许 []:\\/?* 字符，最长 31 字符）."""
    invalid_chars = r'[]:*?/\\'
    for ch in invalid_chars:
        name = name.replace(ch, "_")
    return name[:31] or "Sheet"


def _unique_sheet_name(name: str, existing: set[str]) -> str:
    """确保 sheet 名不重复."""
    if name not in existing:
        return name
    for i in range(2, 100):
        candidate = f"{name[:28]}({i})"
        if candidate not in existing:
            return candidate
    return f"Sheet{len(existing) + 1}"


def _auto_adjust_column_width(
    ws: Any,
    headers: list[str],
    rows: list[list[str]],
    *,
    min_width: int = 8,
    max_width: int = 50,
) -> None:
    """根据内容自动调整列宽."""
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in rows:
            if col_idx <= len(row):
                cell_len = len(str(row[col_idx - 1]))
                max_len = max(max_len, cell_len)
        # 中文字符占 2 个宽度单位
        adjusted_width = min(max(max_len * 1.2 + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


# ──────────────────────────────────────────────
# CLI 入口
# ──────────────────────────────────────────────

def main() -> None:
    """命令行入口."""
    import argparse

    parser = argparse.ArgumentParser(description="Markdown → Excel 转换工具")
    parser.add_argument("input", help="输入的 Markdown 文件路径")
    parser.add_argument("--output", "-o", help="输出的 Excel 文件路径（默认与输入文件同名 .xlsx）")
    parser.add_argument(
        "--mode",
        "-m",
        choices=["tables", "document", "auto"],
        default="auto",
        help="转换模式：tables=仅提取表格, document=全文档, auto=自动判断（默认）",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ 文件不存在: {input_path}")
        sys.exit(1)

    output_path = args.output or input_path.with_suffix(".xlsx")
    md_content = input_path.read_text(encoding="utf-8")

    try:
        result = write_structured_to_excel(md_content, output_path, mode=args.mode)
        print(f"✅ 转换完成: {result}")
    except ValueError as e:
        print(f"❌ {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 转换失败: {type(e).__name__}: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
